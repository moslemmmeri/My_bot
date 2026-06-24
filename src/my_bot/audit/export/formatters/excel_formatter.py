# my_bot_project/src/my_bot/export/formatters/excel_formatter.py
"""
فرمت‌کننده خروجی Excel (Excel Formatter).

این کلاس مسئولیت تبدیل داده‌ها به فرمت Excel (xlsx) را بر عهده دارد.
با استفاده از کتابخانه openpyxl، فایل‌های Excel با قابلیت تنظیم استایل،
رنگ‌بندی و فرمت‌دهی سلول‌ها ایجاد می‌کند.
"""

import io
from typing import List, Dict, Any, Optional, Union
from datetime import datetime

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from my_bot.core.logger.logger_setup import get_logger

logger = get_logger(__name__)


class ExcelFormatter:
    """
    فرمت‌کننده خروجی Excel.

    این کلاس با استفاده از openpyxl، داده‌ها را به‌صورت یک فایل Excel
    با استایل‌دهی مناسب تبدیل می‌کند.

    Attributes:
        header_font: فونت هدرها.
        header_fill: رنگ پس‌زمینه هدرها.
        header_alignment: ترازبندی هدرها.
        cell_alignment: ترازبندی سلول‌ها.
        border: حاشیه سلول‌ها.
        default_sheet_name: نام پیش‌فرض شیت.
        auto_column_width: تنظیم خودکار عرض ستون‌ها (پیش‌فرض True).
    """

    def __init__(
        self,
        auto_column_width: bool = True,
        header_font_size: int = 12,
        header_fill_color: str = "4472C4",
        header_font_color: str = "FFFFFF",
    ) -> None:
        """
        مقداردهی اولیه ExcelFormatter.

        Args:
            auto_column_width: تنظیم خودکار عرض ستون‌ها (پیش‌فرض True).
            header_font_size: اندازه فونت هدرها (پیش‌فرض ۱۲).
            header_fill_color: رنگ پس‌زمینه هدرها (پیش‌فرض آبی).
            header_font_color: رنگ فونت هدرها (پیش‌فرض سفید).
        """
        self.auto_column_width = auto_column_width

        # استایل هدرها
        self.header_font = Font(
            name="B Nazanin",
            size=header_font_size,
            bold=True,
            color=header_font_color,
        )
        self.header_fill = PatternFill(
            start_color=header_fill_color,
            end_color=header_fill_color,
            fill_type="solid",
        )
        self.header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        # استایل سلول‌ها
        self.cell_alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )

        # حاشیه سلول‌ها
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.border = thin_border

        self.default_sheet_name = "Sheet1"

        logger.info(
            f"ExcelFormatter initialized: auto_column_width={auto_column_width}, "
            f"header_fill_color={header_fill_color}"
        )

    async def format_data(
        self,
        data: List[Dict[str, Any]],
        sheet_name: str = "Sheet1",
        include_index: bool = False,
    ) -> bytes:
        """
        تبدیل داده‌ها به فایل Excel.

        Args:
            data: لیست دیکشنری‌های داده.
            sheet_name: نام شیت (پیش‌فرض: "Sheet1").
            include_index: شامل ستون شماره ردیف (پیش‌فرض False).

        Returns:
            bytes: محتوای فایل Excel.

        Raises:
            ValueError: اگر داده‌ها خالی باشند یا فرمت نامعتبر داشته باشند.
        """
        if not data:
            logger.warning("No data provided for Excel export.")
            return self.create_empty_workbook()

        # ایجاد Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # حداکثر ۳۱ کاراکتر

        # استخراج هدرها
        headers = list(data[0].keys())

        # نوشتن هدرها
        start_col = 2 if include_index else 1
        col_offset = 1 if include_index else 0

        # اگر include_index True باشد، ستون شماره ردیف اضافه می‌شود
        if include_index:
            ws.cell(row=1, column=1, value="ردیف")
            self._apply_header_style(ws.cell(row=1, column=1))

        for col_idx, header in enumerate(headers, start=start_col):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_header_style(cell)

        # نوشتن داده‌ها
        for row_idx, row_data in enumerate(data, start=2):
            if include_index:
                ws.cell(row=row_idx, column=1, value=row_idx - 1)
                self._apply_cell_style(ws.cell(row=row_idx, column=1))

            for col_idx, header in enumerate(headers, start=start_col):
                value = row_data.get(header)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                self._apply_cell_style(cell)

        # تنظیم عرض ستون‌ها (در صورت فعال بودن)
        if self.auto_column_width:
            self._auto_fit_columns(ws)

        # ذخیره در bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Excel file created: {len(data)} rows, {len(headers)} columns")
        return output.getvalue()

    async def format_data_with_multiple_sheets(
        self,
        sheets: Dict[str, List[Dict[str, Any]]],
        include_index: bool = False,
    ) -> bytes:
        """
        ایجاد فایل Excel با چندین شیت.

        Args:
            sheets: دیکشنری نگاشت نام شیت به لیست داده‌ها.
            include_index: شامل ستون شماره ردیف (پیش‌فرض False).

        Returns:
            bytes: محتوای فایل Excel.

        Raises:
            ValueError: اگر داده‌ها خالی باشند.
        """
        if not sheets:
            logger.warning("No sheets provided for Excel export.")
            return self.create_empty_workbook()

        wb = Workbook()
        # حذف شیت پیش‌فرض (اولین شیت را با نام مناسب ایجاد می‌کنیم)
        default_sheet = wb.active
        wb.remove(default_sheet)

        for sheet_name, data in sheets.items():
            if not data:
                logger.warning(f"Sheet '{sheet_name}' has no data, skipping.")
                continue

            ws = wb.create_sheet(title=sheet_name[:31])

            # استخراج هدرها از اولین ردیف
            headers = list(data[0].keys())

            # نوشتن هدرها
            start_col = 2 if include_index else 1
            if include_index:
                ws.cell(row=1, column=1, value="ردیف")
                self._apply_header_style(ws.cell(row=1, column=1))

            for col_idx, header in enumerate(headers, start=start_col):
                cell = ws.cell(row=1, column=col_idx, value=header)
                self._apply_header_style(cell)

            # نوشتن داده‌ها
            for row_idx, row_data in enumerate(data, start=2):
                if include_index:
                    ws.cell(row=row_idx, column=1, value=row_idx - 1)
                    self._apply_cell_style(ws.cell(row=row_idx, column=1))

                for col_idx, header in enumerate(headers, start=start_col):
                    value = row_data.get(header)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    self._apply_cell_style(cell)

            # تنظیم عرض ستون‌ها (در صورت فعال بودن)
            if self.auto_column_width:
                self._auto_fit_columns(ws)

        # ذخیره
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Excel file created with {len(sheets)} sheets")
        return output.getvalue()

    def create_empty_workbook(self) -> bytes:
        """
        ایجاد یک فایل Excel خالی.

        Returns:
            bytes: محتوای فایل Excel خالی.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"

        # نوشتن یک پیام ساده در سلول A1
        ws.cell(row=1, column=1, value="هیچ داده‌ای برای خروجی وجود ندارد.")
        self._apply_cell_style(ws.cell(row=1, column=1))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.debug("Empty Excel workbook created.")
        return output.getvalue()

    def _apply_header_style(self, cell) -> None:
        """
        اعمال استایل هدر به یک سلول.

        Args:
            cell: سلول openpyxl.
        """
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.header_alignment
        cell.border = self.border

    def _apply_cell_style(self, cell) -> None:
        """
        اعمال استایل معمولی به یک سلول.

        Args:
            cell: سلول openpyxl.
        """
        cell.alignment = self.cell_alignment
        cell.border = self.border

    def _auto_fit_columns(self, ws: Worksheet) -> None:
        """
        تنظیم خودکار عرض ستون‌ها بر اساس محتوای سلول‌ها.

        Args:
            ws: شیت openpyxl.
        """
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)

            for cell in col:
                if cell.value is not None:
                    try:
                        # محاسبه طول محتوا (تبدیل به رشته)
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                    except Exception:
                        pass

            # محدود کردن عرض برای جلوگیری از ستون‌های خیلی عریض
            adjusted_length = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_length

    def set_auto_column_width(self, enabled: bool) -> None:
        """
        فعال/غیرفعال کردن تنظیم خودکار عرض ستون‌ها.

        Args:
            enabled: True برای فعال، False برای غیرفعال.
        """
        self.auto_column_width = enabled

    def set_header_style(
        self,
        font_size: Optional[int] = None,
        font_color: Optional[str] = None,
        fill_color: Optional[str] = None,
    ) -> None:
        """
        تنظیم استایل هدرها.

        Args:
            font_size: اندازه فونت.
            font_color: رنگ فونت (هگز).
            fill_color: رنگ پس‌زمینه (هگز).
        """
        if font_size is not None:
            self.header_font.size = font_size

        if font_color is not None:
            self.header_font.color = font_color

        if fill_color is not None:
            self.header_fill = PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type="solid",
            )