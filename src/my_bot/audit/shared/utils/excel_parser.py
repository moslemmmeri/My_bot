# my_bot_project/src/my_bot/shared/utils/excel_parser.py
"""
پردازش فایل‌های اکسل (Excel Parser).

این ماژول شامل کلاس `ExcelParser` است که برای خواندن و پردازش
فایل‌های اکسل (xlsx, xls) در سیستم Import استفاده می‌شود.
پشتیبانی از اعتبارسنجی داده‌ها، تبدیل به فرمت‌های مختلف و مدیریت خطاها.
"""

import io
from pathlib import Path
from typing import Optional, List, Dict, Any, Union, Iterator, Tuple
from datetime import datetime

import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from my_bot.core.exceptions.validation_errors import ValidationError
from my_bot.core.logger.logger_setup import get_logger

logger = get_logger(__name__)


class ExcelParser:
    """
    ابزار پردازش فایل‌های اکسل.

    این کلاس با استفاده از openpyxl، فایل‌های اکسل را خوانده و
    داده‌ها را به‌صورت دیکشنری یا لیست بازمی‌گرداند.

    Attributes:
        file_path: مسیر فایل اکسل (اختیاری).
        file_content: محتوای فایل به‌صورت bytes (اختیاری).
        workbook: شیء Workbook openpyxl.
        sheet_names: لیست نام شیت‌ها.
        current_sheet: شیت فعال فعلی.
        header_row: شماره ردیف هدر (پیش‌فرض ۱).
    """

    def __init__(
        self,
        file_path: Optional[Union[str, Path]] = None,
        file_content: Optional[bytes] = None,
        header_row: int = 1,
    ) -> None:
        """
        مقداردهی اولیه ExcelParser.

        Args:
            file_path: مسیر فایل اکسل (اختیاری).
            file_content: محتوای فایل به‌صورت bytes (اختیاری).
            header_row: شماره ردیف هدر (پیش‌فرض ۱).

        Raises:
            ValidationError: اگر فایل معتبر نباشد یا قابل خواندن نباشد.
        """
        self.file_path = Path(file_path) if file_path else None
        self.file_content = file_content
        self.header_row = header_row
        self.workbook: Optional[Workbook] = None
        self.current_sheet: Optional[Worksheet] = None
        self._sheet_names: List[str] = []

        # بارگذاری فایل
        if file_path or file_content:
            self._load_file()

        logger.info(
            f"ExcelParser initialized: file_path={file_path}, "
            f"header_row={header_row}, loaded={self.workbook is not None}"
        )

    def _load_file(self) -> None:
        """
        بارگذاری فایل اکسل از مسیر یا محتوای bytes.

        Raises:
            ValidationError: اگر فایل قابل خواندن نباشد.
        """
        try:
            if self.file_content:
                # بارگذاری از bytes
                file_stream = io.BytesIO(self.file_content)
                self.workbook = load_workbook(file_stream, data_only=True)
            elif self.file_path and self.file_path.exists():
                # بارگذاری از مسیر
                self.workbook = load_workbook(self.file_path, data_only=True)
            else:
                raise ValidationError(
                    message="فایل اکسل یافت نشد.",
                    context={"file_path": str(self.file_path)},
                )

            # دریافت نام شیت‌ها
            self._sheet_names = self.workbook.sheetnames

            # تنظیم شیت فعال (اولین شیت)
            if self._sheet_names:
                self.current_sheet = self.workbook[self._sheet_names[0]]

            logger.info(
                f"Excel file loaded: sheets={self._sheet_names}, "
                f"active_sheet={self.current_sheet.title if self.current_sheet else None}"
            )

        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            raise ValidationError(
                message=f"خطا در بارگذاری فایل اکسل: {str(e)}",
                context={"file_path": str(self.file_path), "error": str(e)},
            )

    def set_sheet(self, sheet_name: Union[str, int]) -> None:
        """
        تنظیم شیت فعال.

        Args:
            sheet_name: نام شیت یا ایندکس (۰ برای اولین شیت).

        Raises:
            ValidationError: اگر شیت وجود نداشته باشد.
        """
        if not self.workbook:
            raise ValidationError(
                message="فایل اکسل بارگذاری نشده است.",
                context={},
            )

        try:
            if isinstance(sheet_name, int):
                if 0 <= sheet_name < len(self._sheet_names):
                    sheet_name = self._sheet_names[sheet_name]
                else:
                    raise ValidationError(
                        message=f"شیت با ایندکس {sheet_name} وجود ندارد.",
                        context={"sheet_index": sheet_name, "total_sheets": len(self._sheet_names)},
                    )

            if sheet_name not in self._sheet_names:
                raise ValidationError(
                    message=f"شیت '{sheet_name}' وجود ندارد.",
                    context={"sheet_name": sheet_name, "available_sheets": self._sheet_names},
                )

            self.current_sheet = self.workbook[sheet_name]
            logger.debug(f"Active sheet set to: {sheet_name}")

        except Exception as e:
            logger.error(f"Error setting sheet: {e}")
            raise ValidationError(
                message=f"خطا در تنظیم شیت: {str(e)}",
                context={"sheet_name": sheet_name},
            )

    def get_sheet_names(self) -> List[str]:
        """
        دریافت لیست نام شیت‌ها.

        Returns:
            List[str]: لیست نام شیت‌ها.
        """
        return self._sheet_names.copy()

    def get_headers(self, sheet_name: Optional[str] = None) -> List[str]:
        """
        دریافت هدرهای شیت.

        Args:
            sheet_name: نام شیت (اختیاری، در صورت None از شیت فعال استفاده می‌شود).

        Returns:
            List[str]: لیست هدرها.

        Raises:
            ValidationError: اگر شیت وجود نداشته باشد یا هدر پیدا نشود.
        """
        sheet = self._get_sheet(sheet_name)

        if self.header_row > sheet.max_row:
            raise ValidationError(
                message=f"ردیف هدر ({self.header_row}) از تعداد ردیف‌ها ({sheet.max_row}) بیشتر است.",
                context={"header_row": self.header_row, "max_row": sheet.max_row},
            )

        headers = []
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=self.header_row, column=col)
            header = self._get_cell_value(cell)
            if header:
                headers.append(str(header))
            else:
                headers.append(f"column_{col}")

        return headers

    def read_rows(
        self,
        sheet_name: Optional[str] = None,
        start_row: Optional[int] = None,
        end_row: Optional[int] = None,
        include_headers: bool = False,
    ) -> List[Dict[str, Any]]:
        """
        خواندن ردیف‌های شیت به‌صورت دیکشنری.

        Args:
            sheet_name: نام شیت (اختیاری).
            start_row: شماره ردیف شروع (اختیاری، پیش‌فرض: header_row + 1).
            end_row: شماره ردیف پایان (اختیاری، پیش‌فرض: آخرین ردیف).
            include_headers: آیا هدرها نیز در خروجی باشند (پیش‌فرض False).

        Returns:
            List[Dict[str, Any]]: لیست دیکشنری‌های داده.

        Raises:
            ValidationError: اگر شیت وجود نداشته باشد یا داده‌ها نامعتبر باشند.
        """
        sheet = self._get_sheet(sheet_name)

        # دریافت هدرها
        headers = self.get_headers(sheet_name)

        # تعیین محدوده ردیف‌ها
        if start_row is None:
            start_row = self.header_row + 1 if not include_headers else 1

        if end_row is None:
            end_row = sheet.max_row

        if start_row > end_row:
            raise ValidationError(
                message="ردیف شروع باید قبل از ردیف پایان باشد.",
                context={"start_row": start_row, "end_row": end_row},
            )

        if end_row > sheet.max_row:
            raise ValidationError(
                message=f"ردیف پایان ({end_row}) از تعداد ردیف‌ها ({sheet.max_row}) بیشتر است.",
                context={"end_row": end_row, "max_row": sheet.max_row},
            )

        # خواندن داده‌ها
        data = []
        for row in range(start_row, end_row + 1):
            row_data = {}
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col)
                row_data[header] = self._get_cell_value(cell)

            data.append(row_data)

        logger.debug(f"Read {len(data)} rows from sheet {sheet.title}")

        return data

    def read_all_sheets(
        self,
        start_row: Optional[int] = None,
        end_row: Optional[int] = None,
        include_headers: bool = False,
    ) -> Dict[str, List[Dict[str, Any]]]:
        """
        خواندن تمام شیت‌های فایل.

        Args:
            start_row: شماره ردیف شروع (اختیاری).
            end_row: شماره ردیف پایان (اختیاری).
            include_headers: آیا هدرها نیز در خروجی باشند (پیش‌فرض False).

        Returns:
            Dict[str, List[Dict[str, Any]]]: دیکشنری نام شیت به لیست داده‌ها.

        Raises:
            ValidationError: اگر فایل بارگذاری نشده باشد.
        """
        if not self.workbook:
            raise ValidationError(
                message="فایل اکسل بارگذاری نشده است.",
                context={},
            )

        result = {}
        for sheet_name in self._sheet_names:
            self.set_sheet(sheet_name)
            result[sheet_name] = self.read_rows(
                sheet_name=sheet_name,
                start_row=start_row,
                end_row=end_row,
                include_headers=include_headers,
            )

        return result

    def read_raw_rows(
        self,
        sheet_name: Optional[str] = None,
        start_row: int = 1,
        end_row: Optional[int] = None,
    ) -> List[List[Any]]:
        """
        خواندن ردیف‌های شیت به‌صورت خام (لیست لیست).

        Args:
            sheet_name: نام شیت (اختیاری).
            start_row: شماره ردیف شروع (پیش‌فرض ۱).
            end_row: شماره ردیف پایان (اختیاری، پیش‌فرض: آخرین ردیف).

        Returns:
            List[List[Any]]: لیست ردیف‌ها.

        Raises:
            ValidationError: اگر شیت وجود نداشته باشد.
        """
        sheet = self._get_sheet(sheet_name)

        if end_row is None:
            end_row = sheet.max_row

        if start_row > end_row:
            raise ValidationError(
                message="ردیف شروع باید قبل از ردیف پایان باشد.",
                context={"start_row": start_row, "end_row": end_row},
            )

        rows = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                row_data.append(self._get_cell_value(cell))
            rows.append(row_data)

        return rows

    def read_column(
        self,
        column: Union[int, str],
        sheet_name: Optional[str] = None,
        start_row: Optional[int] = None,
        end_row: Optional[int] = None,
    ) -> List[Any]:
        """
        خواندن یک ستون خاص از شیت.

        Args:
            column: شماره ستون (از ۱ شروع) یا حرف ستون (مثل 'A').
            sheet_name: نام شیت (اختیاری).
            start_row: شماره ردیف شروع (اختیاری، پیش‌فرض: ۱).
            end_row: شماره ردیف پایان (اختیاری، پیش‌فرض: آخرین ردیف).

        Returns:
            List[Any]: لیست مقادیر ستون.

        Raises:
            ValidationError: اگر شیت وجود نداشته باشد یا ستون نامعتبر باشد.
        """
        sheet = self._get_sheet(sheet_name)

        # تبدیل حرف ستون به شماره
        if isinstance(column, str):
            try:
                col_num = openpyxl.utils.column_index_from_string(column)
            except Exception as e:
                raise ValidationError(
                    message=f"ستون '{column}' نامعتبر است.",
                    context={"column": column, "error": str(e)},
                )
        else:
            col_num = column

        if col_num < 1 or col_num > sheet.max_column:
            raise ValidationError(
                message=f"ستون شماره {col_num} خارج از محدوده است.",
                context={"column": col_num, "max_column": sheet.max_column},
            )

        if start_row is None:
            start_row = 1
        if end_row is None:
            end_row = sheet.max_row

        values = []
        for row in range(start_row, end_row + 1):
            cell = sheet.cell(row=row, column=col_num)
            values.append(self._get_cell_value(cell))

        return values

    def get_cell_value(
        self,
        row: int,
        column: Union[int, str],
        sheet_name: Optional[str] = None,
    ) -> Any:
        """
        دریافت مقدار یک سلول خاص.

        Args:
            row: شماره ردیف.
            column: شماره ستون یا حرف ستون.
            sheet_name: نام شیت (اختیاری).

        Returns:
            Any: مقدار سلول.

        Raises:
            ValidationError: اگر سلول نامعتبر باشد.
        """
        sheet = self._get_sheet(sheet_name)

        # تبدیل حرف ستون به شماره
        if isinstance(column, str):
            try:
                col_num = openpyxl.utils.column_index_from_string(column)
            except Exception as e:
                raise ValidationError(
                    message=f"ستون '{column}' نامعتبر است.",
                    context={"column": column, "error": str(e)},
                )
        else:
            col_num = column

        if row < 1 or row > sheet.max_row:
            raise ValidationError(
                message=f"ردیف {row} خارج از محدوده است.",
                context={"row": row, "max_row": sheet.max_row},
            )

        if col_num < 1 or col_num > sheet.max_column:
            raise ValidationError(
                message=f"ستون {col_num} خارج از محدوده است.",
                context={"column": col_num, "max_column": sheet.max_column},
            )

        cell = sheet.cell(row=row, column=col_num)
        return self._get_cell_value(cell)

    def get_sheet_dimensions(self, sheet_name: Optional[str] = None) -> Dict[str, int]:
        """
        دریافت ابعاد شیت (تعداد ردیف‌ها و ستون‌ها).

        Args:
            sheet_name: نام شیت (اختیاری).

        Returns:
            Dict[str, int]: دیکشنری شامل rows و columns.

        Raises:
            ValidationError: اگر شیت وجود نداشته باشد.
        """
        sheet = self._get_sheet(sheet_name)
        return {
            "rows": sheet.max_row,
            "columns": sheet.max_column,
        }

    def validate_headers(
        self,
        required_headers: List[str],
        sheet_name: Optional[str] = None,
    ) -> Dict[str, bool]:
        """
        اعتبارسنجی وجود هدرهای مورد نیاز در شیت.

        Args:
            required_headers: لیست هدرهای مورد نیاز.
            sheet_name: نام شیت (اختیاری).

        Returns:
            Dict[str, bool]: دیکشنری نام هدر به وجود/عدم وجود.

        Raises:
            ValidationError: اگر شیت وجود نداشته باشد.
        """
        headers = self.get_headers(sheet_name)
        result = {}

        for required in required_headers:
            result[required] = required in headers

        return result

    def validate_headers_raise(
        self,
        required_headers: List[str],
        sheet_name: Optional[str] = None,
    ) -> None:
        """
        اعتبارسنجی هدرهای مورد نیاز و پرتاب خطا در صورت عدم وجود.

        Args:
            required_headers: لیست هدرهای مورد نیاز.
            sheet_name: نام شیت (اختیاری).

        Raises:
            ValidationError: اگر هدر مورد نیاز وجود نداشته باشد.
        """
        result = self.validate_headers(required_headers, sheet_name)
        missing = [h for h, exists in result.items() if not exists]

        if missing:
            raise ValidationError(
                message=f"هدرهای زیر در فایل وجود ندارند: {', '.join(missing)}",
                context={"missing_headers": missing, "required_headers": required_headers},
            )

    def get_data_by_headers(
        self,
        required_headers: List[str],
        sheet_name: Optional[str] = None,
        start_row: Optional[int] = None,
        end_row: Optional[int] = None,
    ) -> List[Dict[str, Any]]:
        """
        خواندن داده‌ها با تأیید وجود هدرهای مورد نیاز.

        Args:
            required_headers: لیست هدرهای مورد نیاز.
            sheet_name: نام شیت (اختیاری).
            start_row: شماره ردیف شروع (اختیاری).
            end_row: شماره ردیف پایان (اختیاری).

        Returns:
            List[Dict[str, Any]]: لیست داده‌ها.

        Raises:
            ValidationError: اگر هدرهای مورد نیاز وجود نداشته باشند.
        """
        self.validate_headers_raise(required_headers, sheet_name)
        return self.read_rows(sheet_name, start_row, end_row)

    def get_data_by_headers_with_validation(
        self,
        required_headers: List[str],
        validators: Dict[str, callable],
        sheet_name: Optional[str] = None,
        start_row: Optional[int] = None,
        end_row: Optional[int] = None,
    ) -> List[Dict[str, Any]]:
        """
        خواندن داده‌ها با اعتبارسنجی مقادیر.

        Args:
            required_headers: لیست هدرهای مورد نیاز.
            validators: دیکشنری نگاشت هدر به تابع اعتبارسنجی.
            sheet_name: نام شیت (اختیاری).
            start_row: شماره ردیف شروع (اختیاری).
            end_row: شماره ردیف پایان (اختیاری).

        Returns:
            List[Dict[str, Any]]: لیست داده‌های معتبر.

        Raises:
            ValidationError: اگر داده‌ها نامعتبر باشند یا هدرها وجود نداشته باشند.
        """
        data = self.get_data_by_headers(required_headers, sheet_name, start_row, end_row)

        validated_data = []
        errors = []

        for idx, row in enumerate(data, start=1):
            row_errors = []
            for header, value in row.items():
                if header in validators:
                    try:
                        validators[header](value)
                    except Exception as e:
                        row_errors.append({
                            "row": idx + (start_row or self.header_row + 1),
                            "header": header,
                            "value": value,
                            "error": str(e),
                        })

            if row_errors:
                errors.extend(row_errors)
            else:
                validated_data.append(row)

        if errors:
            raise ValidationError(
                message=f"خطاهای اعتبارسنجی در {len(errors)} مورد.",
                context={"errors": errors},
            )

        return validated_data

    def _get_sheet(self, sheet_name: Optional[str] = None) -> Worksheet:
        """
        دریافت شیت مورد نظر.

        Args:
            sheet_name: نام شیت (اختیاری).

        Returns:
            Worksheet: شیت.

        Raises:
            ValidationError: اگر شیت وجود نداشته باشد یا فایل بارگذاری نشده باشد.
        """
        if not self.workbook:
            raise ValidationError(
                message="فایل اکسل بارگذاری نشده است.",
                context={},
            )

        if sheet_name:
            if sheet_name not in self._sheet_names:
                raise ValidationError(
                    message=f"شیت '{sheet_name}' وجود ندارد.",
                    context={"sheet_name": sheet_name, "available_sheets": self._sheet_names},
                )
            return self.workbook[sheet_name]

        if self.current_sheet:
            return self.current_sheet

        # اگر شیت فعال وجود ندارد، اولین شیت را انتخاب کن
        if self._sheet_names:
            self.current_sheet = self.workbook[self._sheet_names[0]]
            return self.current_sheet

        raise ValidationError(
            message="هیچ شیتی در فایل وجود ندارد.",
            context={},
        )

    def _get_cell_value(self, cell: Cell) -> Any:
        """
        دریافت مقدار سلول با مدیریت انواع داده‌ها.

        Args:
            cell: سلول openpyxl.

        Returns:
            Any: مقدار سلول.
        """
        value = cell.value

        # مدیریت None
        if value is None:
            return None

        # مدیریت datetime
        if isinstance(value, datetime):
            return value.isoformat()

        # مدیریت فرمول‌ها (در صورت استفاده از data_only=False)
        if isinstance(value, str) and value.startswith("="):
            return f"FORMULA:{value}"

        return value

    def to_dict(self) -> Dict[str, Any]:
        """
        تبدیل کل فایل به دیکشنری (برای سریال‌سازی).

        Returns:
            Dict[str, Any]: دیکشنری شامل اطلاعات فایل.
        """
        if not self.workbook:
            return {"error": "File not loaded"}

        return {
            "file_path": str(self.file_path) if self.file_path else None,
            "sheets": self._sheet_names,
            "active_sheet": self.current_sheet.title if self.current_sheet else None,
            "data": self.read_all_sheets(),
        }

    def close(self) -> None:
        """
        بستن فایل و آزادسازی منابع.
        """
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.current_sheet = None
            self._sheet_names = []
            logger.info("Excel file closed.")


# ==========================================
# توابع کمکی
# ==========================================

def create_excel_from_dict(
    data: List[Dict[str, Any]],
    sheet_name: str = "Sheet1",
) -> bytes:
    """
    ایجاد فایل اکسل از داده‌های دیکشنری.

    Args:
        data: لیست دیکشنری‌های داده.
        sheet_name: نام شیت (پیش‌فرض: "Sheet1").

    Returns:
        bytes: محتوای فایل اکسل.

    Raises:
        ValidationError: اگر داده‌ها نامعتبر باشند.
    """
    if not data:
        raise ValidationError(
            message="داده‌ها نمی‌توانند خالی باشند.",
            context={},
        )

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # دریافت هدرها از اولین ردیف
        headers = list(data[0].keys())

        # نوشتن هدرها
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # نوشتن داده‌ها
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))

        # ذخیره در bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Excel file created: {len(data)} rows, {len(headers)} columns")

        return output.getvalue()

    except Exception as e:
        logger.error(f"Error creating Excel file: {e}")
        raise ValidationError(
            message=f"خطا در ایجاد فایل اکسل: {str(e)}",
            context={"error": str(e)},
        )