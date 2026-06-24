# my_bot_project/src/my_bot/bulk_import/parsers/excel_reader.py
"""
خواننده فایل‌های اکسل (Excel Reader).

این کلاس با استفاده از openpyxl، فایل‌های اکسل (xlsx, xls) را برای
واردات انبوه داده‌ها خوانده و به فرمت مناسب تبدیل می‌کند.
"""

from pathlib import Path
from typing import Optional, List, Dict, Any, Union, Iterator
from io import BytesIO

import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from my_bot.core.exceptions.validation_errors import ValidationError
from my_bot.core.logger.logger_setup import get_logger

logger = get_logger(__name__)


class ExcelReader:
    """
    خواننده فایل‌های اکسل برای واردات انبوه.

    این کلاس با استفاده از openpyxl، فایل‌های اکسل را خوانده و
    داده‌ها را به‌صورت لیست دیکشنری یا لیست لیست بازمی‌گرداند.

    Attributes:
        file_path: مسیر فایل اکسل (اختیاری).
        file_content: محتوای فایل به‌صورت bytes (اختیاری).
        header_row: شماره ردیف هدر (پیش‌فرض ۱).
        sheet_name: نام شیت (در صورت عدم مشخص بودن، شیت فعال استفاده می‌شود).
        workbook: شیء Workbook openpyxl.
        worksheet: شیت فعال.
        headers: لیست هدرها.
        _is_loaded: وضعیت بارگذاری فایل.
    """

    def __init__(
        self,
        file_path: Optional[Union[str, Path]] = None,
        file_content: Optional[bytes] = None,
        header_row: int = 1,
        sheet_name: Optional[str] = None,
    ) -> None:
        """
        مقداردهی اولیه ExcelReader.

        Args:
            file_path: مسیر فایل اکسل (اختیاری).
            file_content: محتوای فایل به‌صورت bytes (اختیاری).
            header_row: شماره ردیف هدر (پیش‌فرض ۱).
            sheet_name: نام شیت (اختیاری).

        Raises:
            ValidationError: اگر فایل معتبر نباشد یا قابل خواندن نباشد.
        """
        self.file_path = Path(file_path) if file_path else None
        self.file_content = file_content
        self.header_row = header_row
        self.sheet_name = sheet_name
        self.workbook: Optional[openpyxl.Workbook] = None
        self.worksheet: Optional[Worksheet] = None
        self.headers: List[str] = []
        self._is_loaded = False

        # بارگذاری فایل در صورت وجود
        if file_path or file_content:
            self._load_file()

        logger.info(
            f"ExcelReader initialized: file_path={file_path}, "
            f"header_row={header_row}, sheet_name={sheet_name}"
        )

    def _load_file(self) -> None:
        """
        بارگذاری فایل اکسل از مسیر یا محتوای bytes.

        Raises:
            ValidationError: اگر فایل قابل خواندن نباشد.
        """
        try:
            if self.file_content:
                # بارگذاری از bytes
                file_stream = BytesIO(self.file_content)
                self.workbook = load_workbook(file_stream, data_only=True)
                logger.debug("Excel file loaded from bytes")
            elif self.file_path and self.file_path.exists():
                # بارگذاری از مسیر
                self.workbook = load_workbook(self.file_path, data_only=True)
                logger.debug(f"Excel file loaded from path: {self.file_path}")
            else:
                raise ValidationError(
                    message="فایل اکسل یافت نشد.",
                    context={"file_path": str(self.file_path)},
                )

            # انتخاب شیت
            self._select_worksheet()

            # خواندن هدرها
            self._read_headers()

            self._is_loaded = True
            logger.info(
                f"Excel file loaded: sheets={self.workbook.sheetnames}, "
                f"headers={len(self.headers)}, rows={self.worksheet.max_row if self.worksheet else 0}"
            )

        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            raise ValidationError(
                message=f"خطا در بارگذاری فایل اکسل: {str(e)}",
                context={"file_path": str(self.file_path), "error": str(e)},
            )

    def _select_worksheet(self) -> None:
        """
        انتخاب شیت فعال بر اساس نام یا اولین شیت.

        Raises:
            ValidationError: اگر شیت وجود نداشته باشد.
        """
        if not self.workbook:
            raise ValidationError(
                message="فایل اکسل بارگذاری نشده است.",
                context={},
            )

        sheet_names = self.workbook.sheetnames
        if not sheet_names:
            raise ValidationError(
                message="فایل اکسل هیچ شیتی ندارد.",
                context={},
            )

        if self.sheet_name:
            if self.sheet_name not in sheet_names:
                raise ValidationError(
                    message=f"شیت '{self.sheet_name}' در فایل وجود ندارد.",
                    context={"sheet_name": self.sheet_name, "available_sheets": sheet_names},
                )
            self.worksheet = self.workbook[self.sheet_name]
        else:
            # استفاده از اولین شیت
            self.worksheet = self.workbook[sheet_names[0]]

        logger.debug(f"Selected worksheet: {self.worksheet.title}")

    def _read_headers(self) -> None:
        """
        خواندن هدرها از ردیف مشخص‌شده.

        Raises:
            ValidationError: اگر ردیف هدر خالی باشد یا نامعتبر باشد.
        """
        if not self.worksheet:
            raise ValidationError(
                message="شیت انتخاب نشده است.",
                context={},
            )

        if self.header_row > self.worksheet.max_row:
            raise ValidationError(
                message=f"ردیف هدر ({self.header_row}) از تعداد ردیف‌ها ({self.worksheet.max_row}) بیشتر است.",
                context={"header_row": self.header_row, "max_row": self.worksheet.max_row},
            )

        headers = []
        for col in range(1, self.worksheet.max_column + 1):
            cell = self.worksheet.cell(row=self.header_row, column=col)
            header = self._get_cell_value(cell)
            if header:
                headers.append(str(header).strip())
            else:
                # اگر هدر خالی بود، یک نام پیش‌فرض ایجاد می‌کنیم
                headers.append(f"column_{col}")

        if not headers or all(h == "" for h in headers):
            raise ValidationError(
                message="ردیف هدر خالی است. لطفاً یک ردیف معتبر برای هدر انتخاب کنید.",
                context={"header_row": self.header_row},
            )

        self.headers = headers
        logger.debug(f"Headers read: {len(self.headers)} columns")

    def _get_cell_value(self, cell: Cell) -> Any:
        """
        دریافت مقدار سلول با مدیریت انواع داده‌ها.

        Args:
            cell: سلول openpyxl.

        Returns:
            Any: مقدار سلول.
        """
        value = cell.value

        if value is None:
            return None

        # مدیریت datetime
        if isinstance(value, (openpyxl.utils.datetime.Cell, openpyxl.utils.datetime.datetime)):
            return value.isoformat()

        # مدیریت فرمول‌ها (در صورت استفاده از data_only=False)
        if isinstance(value, str) and value.startswith("="):
            return f"FORMULA:{value}"

        return value

    def read_rows(
        self,
        start_row: Optional[int] = None,
        end_row: Optional[int] = None,
        include_headers: bool = False,
        as_dict: bool = True,
    ) -> Union[List[Dict[str, Any]], List[List[Any]]]:
        """
        خواندن ردیف‌های شیت به‌صورت دیکشنری یا لیست.

        Args:
            start_row: شماره ردیف شروع (اختیاری، پیش‌فرض: header_row + 1).
            end_row: شماره ردیف پایان (اختیاری، پیش‌فرض: آخرین ردیف).
            include_headers: آیا هدرها نیز در خروجی باشند (پیش‌فرض False).
            as_dict: خروجی به‌صورت دیکشنری (پیش‌فرض True) یا لیست.

        Returns:
            لیست دیکشنری‌ها یا لیست لیست‌های داده.

        Raises:
            ValidationError: اگر شیت انتخاب نشده باشد یا ردیف‌ها نامعتبر باشند.
        """
        if not self._is_loaded or not self.worksheet:
            raise ValidationError(
                message="فایل اکسل بارگذاری نشده است.",
                context={},
            )

        # تعیین محدوده ردیف‌ها
        if start_row is None:
            start_row = self.header_row + 1 if not include_headers else 1

        if end_row is None:
            end_row = self.worksheet.max_row

        if start_row > end_row:
            raise ValidationError(
                message="ردیف شروع باید قبل از ردیف پایان باشد.",
                context={"start_row": start_row, "end_row": end_row},
            )

        if end_row > self.worksheet.max_row:
            raise ValidationError(
                message=f"ردیف پایان ({end_row}) از تعداد ردیف‌ها ({self.worksheet.max_row}) بیشتر است.",
                context={"end_row": end_row, "max_row": self.worksheet.max_row},
            )

        data = []

        for row_num in range(start_row, end_row + 1):
            row_data = []
            for col_num in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=row_num, column=col_num)
                row_data.append(self._get_cell_value(cell))

            if as_dict:
                # ساخت دیکشنری با استفاده از هدرها
                # اگر include_headers=True، هدرها نیز به‌عنوان داده در نظر گرفته می‌شوند
                row_dict = {}
                for idx, header in enumerate(self.headers):
                    if idx < len(row_data):
                        row_dict[header] = row_data[idx]
                    else:
                        row_dict[header] = None
                data.append(row_dict)
            else:
                data.append(row_data)

        logger.debug(f"Read {len(data)} rows from {start_row} to {end_row}")
        return data

    def read_column(
        self,
        column: Union[int, str],
        start_row: Optional[int] = None,
        end_row: Optional[int] = None,
    ) -> List[Any]:
        """
        خواندن یک ستون خاص از شیت.

        Args:
            column: شماره ستون (از ۱ شروع) یا حرف ستون (مثل 'A').
            start_row: شماره ردیف شروع (اختیاری، پیش‌فرض: ۱).
            end_row: شماره ردیف پایان (اختیاری، پیش‌فرض: آخرین ردیف).

        Returns:
            List[Any]: لیست مقادیر ستون.

        Raises:
            ValidationError: اگر شیت انتخاب نشده باشد یا ستون نامعتبر باشد.
        """
        if not self._is_loaded or not self.worksheet:
            raise ValidationError(
                message="فایل اکسل بارگذاری نشده است.",
                context={},
            )

        # تبدیل حرف ستون به شماره
        if isinstance(column, str):
            try:
                col_num = openpyxl.utils.column_index_from_string(column)
            except Exception as e:
                raise ValidationError(
                    message=f"ستون '{column}' نامعتبر است.",
                    context={"column": column, "error": str(e)},
                )
        else:
            col_num = column

        if col_num < 1 or col_num > self.worksheet.max_column:
            raise ValidationError(
                message=f"ستون شماره {col_num} خارج از محدوده است.",
                context={"column": col_num, "max_column": self.worksheet.max_column},
            )

        if start_row is None:
            start_row = 1
        if end_row is None:
            end_row = self.worksheet.max_row

        values = []
        for row in range(start_row, end_row + 1):
            cell = self.worksheet.cell(row=row, column=col_num)
            values.append(self._get_cell_value(cell))

        return values

    def get_cell_value(self, row: int, column: Union[int, str]) -> Any:
        """
        دریافت مقدار یک سلول خاص.

        Args:
            row: شماره ردیف.
            column: شماره ستون یا حرف ستون.

        Returns:
            Any: مقدار سلول.

        Raises:
            ValidationError: اگر سلول نامعتبر باشد.
        """
        if not self._is_loaded or not self.worksheet:
            raise ValidationError(
                message="فایل اکسل بارگذاری نشده است.",
                context={},
            )

        if isinstance(column, str):
            try:
                col_num = openpyxl.utils.column_index_from_string(column)
            except Exception as e:
                raise ValidationError(
                    message=f"ستون '{column}' نامعتبر است.",
                    context={"column": column, "error": str(e)},
                )
        else:
            col_num = column

        if row < 1 or row > self.worksheet.max_row:
            raise ValidationError(
                message=f"ردیف {row} خارج از محدوده است.",
                context={"row": row, "max_row": self.worksheet.max_row},
            )

        if col_num < 1 or col_num > self.worksheet.max_column:
            raise ValidationError(
                message=f"ستون {col_num} خارج از محدوده است.",
                context={"column": col_num, "max_column": self.worksheet.max_column},
            )

        cell = self.worksheet.cell(row=row, column=col_num)
        return self._get_cell_value(cell)

    def get_dimensions(self) -> Dict[str, int]:
        """
        دریافت ابعاد شیت (تعداد ردیف‌ها و ستون‌ها).

        Returns:
            Dict[str, int]: دیکشنری شامل rows و columns.

        Raises:
            ValidationError: اگر شیت انتخاب نشده باشد.
        """
        if not self._is_loaded or not self.worksheet:
            raise ValidationError(
                message="فایل اکسل بارگذاری نشده است.",
                context={},
            )

        return {
            "rows": self.worksheet.max_row,
            "columns": self.worksheet.max_column,
        }

    def get_sheet_names(self) -> List[str]:
        """
        دریافت لیست نام شیت‌ها.

        Returns:
            List[str]: لیست نام شیت‌ها.

        Raises:
            ValidationError: اگر فایل بارگذاری نشده باشد.
        """
        if not self._is_loaded or not self.workbook:
            raise ValidationError(
                message="فایل اکسل بارگذاری نشده است.",
                context={},
            )

        return self.workbook.sheetnames.copy()

    def set_sheet(self, sheet_name: str) -> None:
        """
        تغییر شیت فعال.

        Args:
            sheet_name: نام شیت جدید.

        Raises:
            ValidationError: اگر شیت وجود نداشته باشد.
        """
        if not self._is_loaded or not self.workbook:
            raise ValidationError(
                message="فایل اکسل بارگذاری نشده است.",
                context={},
            )

        if sheet_name not in self.workbook.sheetnames:
            raise ValidationError(
                message=f"شیت '{sheet_name}' در فایل وجود ندارد.",
                context={"sheet_name": sheet_name, "available_sheets": self.workbook.sheetnames},
            )

        self.worksheet = self.workbook[sheet_name]
        self._read_headers()
        logger.debug(f"Sheet changed to: {sheet_name}")

    def validate_headers(
        self,
        required_headers: List[str],
        case_sensitive: bool = False,
    ) -> Dict[str, bool]:
        """
        اعتبارسنجی وجود هدرهای مورد نیاز.

        Args:
            required_headers: لیست هدرهای مورد نیاز.
            case_sensitive: آیا تطابق حروف بزرگ/کوچک مهم است (پیش‌فرض False).

        Returns:
            Dict[str, bool]: دیکشنری نام هدر به وجود/عدم وجود.

        Raises:
            ValidationError: اگر شیت انتخاب نشده باشد.
        """
        if not self._is_loaded or not self.worksheet:
            raise ValidationError(
                message="فایل اکسل بارگذاری نشده است.",
                context={},
            )

        headers_lower = [h.lower() for h in self.headers]
        result = {}

        for required in required_headers:
            if case_sensitive:
                result[required] = required in self.headers
            else:
                result[required] = required.lower() in headers_lower

        return result

    def validate_headers_raise(
        self,
        required_headers: List[str],
        case_sensitive: bool = False,
    ) -> None:
        """
        اعتبارسنجی هدرهای مورد نیاز و پرتاب خطا در صورت عدم وجود.

        Args:
            required_headers: لیست هدرهای مورد نیاز.
            case_sensitive: آیا تطابق حروف بزرگ/کوچک مهم است (پیش‌فرض False).

        Raises:
            ValidationError: اگر هدر مورد نیاز وجود نداشته باشد.
        """
        result = self.validate_headers(required_headers, case_sensitive)
        missing = [h for h, exists in result.items() if not exists]

        if missing:
            raise ValidationError(
                message=f"هدرهای زیر در فایل وجود ندارند: {', '.join(missing)}",
                context={"missing_headers": missing, "required_headers": required_headers},
            )

    def get_non_empty_rows_count(self, start_row: Optional[int] = None) -> int:
        """
        شمارش تعداد ردیف‌های غیرخالی.

        Args:
            start_row: شماره ردیف شروع (اختیاری، پیش‌فرض: header_row + 1).

        Returns:
            int: تعداد ردیف‌های غیرخالی.

        Raises:
            ValidationError: اگر شیت انتخاب نشده باشد.
        """
        if not self._is_loaded or not self.worksheet:
            raise ValidationError(
                message="فایل اکسل بارگذاری نشده است.",
                context={},
            )

        if start_row is None:
            start_row = self.header_row + 1

        count = 0
        for row in range(start_row, self.worksheet.max_row + 1):
            is_empty = True
            for col in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=row, column=col)
                if cell.value is not None:
                    is_empty = False
                    break
            if not is_empty:
                count += 1

        return count

    def get_sheet_summary(self) -> Dict[str, Any]:
        """
        دریافت خلاصه اطلاعات شیت.

        Returns:
            Dict[str, Any]: خلاصه شیت شامل تعداد ردیف‌ها، ستون‌ها، هدرها و ...
        """
        if not self._is_loaded or not self.worksheet:
            return {
                "is_loaded": False,
                "message": "فایل اکسل بارگذاری نشده است.",
            }

        return {
            "is_loaded": True,
            "sheet_name": self.worksheet.title,
            "total_rows": self.worksheet.max_row,
            "total_columns": self.worksheet.max_column,
            "headers": self.headers,
            "header_row": self.header_row,
            "non_empty_rows": self.get_non_empty_rows_count(),
            "columns": [
                {
                    "index": i + 1,
                    "letter": get_column_letter(i + 1),
                    "header": self.headers[i] if i < len(self.headers) else f"col_{i + 1}",
                }
                for i in range(self.worksheet.max_column)
            ],
        }

    def close(self) -> None:
        """
        بستن فایل و آزادسازی منابع.
        """
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.worksheet = None
            self.headers = []
            self._is_loaded = False
            logger.info("Excel file closed.")

    def __enter__(self):
        """پشتیبانی از Context Manager."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """خروج از Context Manager."""
        self.close()

    def __iter__(self) -> Iterator[Dict[str, Any]]:
        """
        Iterator برای پیمایش ردیف‌های داده.

        Yields:
            Dict[str, Any]: هر ردیف به‌صورت دیکشنری.

        Raises:
            ValidationError: اگر فایل بارگذاری نشده باشد.
        """
        if not self._is_loaded or not self.worksheet:
            raise ValidationError(
                message="فایل اکسل بارگذاری نشده است.",
                context={},
            )

        for row_num in range(self.header_row + 1, self.worksheet.max_row + 1):
            row_data = {}
            for col_num in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=row_num, column=col_num)
                header = self.headers[col_num - 1] if col_num - 1 < len(self.headers) else f"col_{col_num}"
                row_data[header] = self._get_cell_value(cell)

            # بررسی خالی بودن ردیف
            if all(v is None for v in row_data.values()):
                break

            yield row_data